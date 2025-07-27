#!/usr/bin/env python3
"""
Script sederhana untuk menambahkan data perusahaan ke Excel
"""

from openpyxl import load_workbook
from datetime import datetime
import pandas as pd

def add_company_data():
    """Menambahkan data perusahaan ke Excel"""
    
    # Data perusahaan yang akan ditambahkan
    companies = [
        {
            'nama': 'PT Astra International Tbk',
            'sektor': 'Otomotif',
            'website': 'https://www.astra.co.id',
            'social_media': 'Instagram: @astra.co.id\nLinkedIn: Astra International',
            'kontak': '+62-21-508-8888'
        },
        {
            'nama': 'PT Bank Central Asia Tbk',
            'sektor': 'Perbankan',
            'website': 'https://www.bca.co.id',
            'social_media': 'Instagram: @bca.co.id\nLinkedIn: Bank Central Asia',
            'kontak': '+62-21-2358-8000'
        },
        {
            'nama': 'PT Telekomunikasi Indonesia Tbk',
            'sektor': 'Telekomunikasi',
            'website': 'https://www.telkom.co.id',
            'social_media': 'Instagram: @telkomindonesia\nLinkedIn: Telkom Indonesia',
            'kontak': '+62-21-521-1234'
        },
        {
            'nama': 'PT Bank Rakyat Indonesia Tbk',
            'sektor': 'Perbankan',
            'website': 'https://www.bri.co.id',
            'social_media': 'Instagram: @bank.bri\nLinkedIn: Bank Rakyat Indonesia',
            'kontak': '+62-21-251-0251'
        },
        {
            'nama': 'PT Bank Mandiri Tbk',
            'sektor': 'Perbankan',
            'website': 'https://www.bankmandiri.co.id',
            'social_media': 'Instagram: @bankmandiri\nLinkedIn: Bank Mandiri',
            'kontak': '+62-21-524-5000'
        }
    ]
    
    try:
        # Load Excel file
        print("📊 Membuka file Excel...")
        wb = load_workbook('Monthly Plan NEXT IT 2024.xlsx')
        ws = wb['Data Pipeline']
        
        # Get today's date
        today = datetime.now().strftime('%Y-%m-%d')
        
        print(f"📅 Tanggal: {today}")
        print(f"📋 Akan menambahkan {len(companies)} perusahaan...")
        
        # Add each company
        for i, company in enumerate(companies, 1):
            row = [
                '',  # No (auto)
                today,  # Tanggal Input
                '',  # Pipeline ID (auto)
                company['nama'],
                company['sektor'],
                company['website'],
                company['social_media'],
                company['kontak'],
                'Manual Input'  # Source
            ]
            
            ws.append(row)
            print(f"✅ {i}. {company['nama']}")
        
        # Save the file
        wb.save('Monthly Plan NEXT IT 2024.xlsx')
        print(f"\n🎉 Berhasil menambahkan {len(companies)} data perusahaan!")
        print("📁 File tersimpan: Monthly Plan NEXT IT 2024.xlsx")
        
    except FileNotFoundError:
        print("❌ File 'Monthly Plan NEXT IT 2024.xlsx' tidak ditemukan!")
        print("📋 Pastikan file ada di folder yang sama")
    except Exception as e:
        print(f"❌ Error: {str(e)}")

def add_from_csv():
    """Menambahkan data dari CSV ke Excel"""
    
    try:
        # Read CSV
        print("📊 Membaca file CSV...")
        df = pd.read_csv('daftar_perusahaan_idx.csv')
        print(f"📋 Total data di CSV: {len(df)}")
        
        # Load Excel
        wb = load_workbook('Monthly Plan NEXT IT 2024.xlsx')
        ws = wb['Data Pipeline']
        
        today = datetime.now().strftime('%Y-%m-%d')
        
        # Add first 20 companies from CSV
        count = 0
        for idx, row in df.head(20).iterrows():
            nama = row['Nama Perusahaan']
            if isinstance(nama, str) and 'BEI:' in nama:
                nama = nama.split('BEI:')[1].strip()
            
            excel_row = [
                '',  # No
                today,  # Tanggal
                '',  # Pipeline ID
                nama,
                row.get('Sektor Bisnis', ''),
                row.get('Website', ''),
                row.get('Social Media', ''),
                row.get('Kontak', ''),
                'IDX CSV'
            ]
            
            ws.append(excel_row)
            count += 1
            print(f"✅ {count}. {nama}")
        
        # Save
        wb.save('Monthly Plan NEXT IT 2024.xlsx')
        print(f"\n🎉 Berhasil menambahkan {count} data dari CSV!")
        
    except FileNotFoundError:
        print("❌ File CSV tidak ditemukan!")
    except Exception as e:
        print(f"❌ Error: {str(e)}")

def main():
    """Main function"""
    print("🚀 SIMPLE ADD TO EXCEL")
    print("=" * 30)
    
    print("\n📋 Pilihan:")
    print("1. Tambah data sample (5 perusahaan)")
    print("2. Tambah dari CSV (20 perusahaan)")
    
    try:
        choice = input("\n❓ Pilih (1/2): ").strip()
        
        if choice == '1':
            add_company_data()
        elif choice == '2':
            add_from_csv()
        else:
            print("❌ Pilihan tidak valid!")
            
    except KeyboardInterrupt:
        print("\n❌ Proses dibatalkan.")
    except Exception as e:
        print(f"❌ Error: {str(e)}")

if __name__ == "__main__":
    main() 